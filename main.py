import xlwt
import xlrd
from xlutils.copy import copy
import xlsxwriter
import collections
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm, trange
from datetime import date, datetime
def xuly(namedata, nameOT,namenhanvien, text_nam, text_thang):
    def myround(x, base=0.5):
        return base * round(float(x) / base)
    # =========================== convert OT =====================
    print("Chuan bi du lieu")
    dataOT = xlrd.open_workbook(nameOT)
    ot = dataOT.sheet_by_index(0)
    ot_convert = xlsxwriter.Workbook('../cham-cong/convert/OT_convert.xlsx')
    add_sheet = ot_convert.add_worksheet()
    for i in range(ot.nrows-3):
        id = ot.cell_value(i+3, 0)
        date = ot.cell_value(i+3, 4)
        date1 =ot.cell_value(i+3, 5)
        x =(datetime.strptime(date,"%Y-%m-%d %H:%M:%S"))
        y =(datetime.strptime(date1,"%Y-%m-%d %H:%M:%S"))
        timeOT = y - x
        hh, mm , ss = map(int, str(timeOT).split(':'))
        ot3 = hh + mm/60
        d = x.strftime("%Y-%m-%d")
        add_sheet.write(i,0, str(id))
        add_sheet.write(i,1, ot3)
        
        add_sheet.write(i,2, d)
    ot_convert.close()

    # =========================== Xoá data =====================

    baocao_del = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao_del = baocao_del.sheet_by_index(0)

    all_rows_baocao_del = []
    for row in tqdm(range(data_baocao_del.nrows)):
        curr_row = []
        for col in range(data_baocao_del.ncols):
            curr_row.append(data_baocao_del.cell_value(row, col))
        all_rows_baocao_del.append(curr_row)

    delete_baocao = xlsxwriter.Workbook('../cham-cong/convert/baocao.xlsx')
    delete = delete_baocao.add_worksheet()

    for row in range(len(all_rows_baocao_del)):
        for col in range(len(all_rows_baocao_del[0])):
            delete.write(row, col, "")
    delete_baocao.close()

    print("Hop nhat ca trong ngay")
    chamcong = xlrd.open_workbook(namedata)
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)
    for ID in range(data.nrows -3):
        if(data.cell_value(ID+2, 0) == data.cell_value(ID+3, 0) and data.cell_value(ID+2, 3) == data.cell_value(ID+3, 3)):
            if (data.cell_value(ID+2, 26) == '' and data.cell_value(ID+3, 26) != ''):
                lateIn = float(data.cell_value(ID+3, 26))
            elif data.cell_value(ID+2, 26) != '' and data.cell_value(ID+3, 26) == '':
                lateIn = float(data.cell_value(ID+2, 26))
            elif data.cell_value(ID+3, 26) == '' and data.cell_value(ID+3, 26) == '':
                lateIn = 0
            else: 
                lateIn = float(data.cell_value(ID+2, 26)) + float(data.cell_value(ID+3, 26))

            if (data.cell_value(ID+2, 27) == '' and data.cell_value(ID+3, 27) != ''):
                earlyOut = float(data.cell_value(ID+3, 27))
            elif data.cell_value(ID+2, 27) != '' and data.cell_value(ID+3, 27) == '':
                earlyOut = float(data.cell_value(ID+2, 27))
            elif data.cell_value(ID+3, 27) == '' and data.cell_value(ID+3, 27) == '':
                earlyOut = 0
            else: 
                earlyOut = float(data.cell_value(ID+2, 27)) + float(data.cell_value(ID+3, 27))    

            if (data.cell_value(ID+2, 28) == '' and data.cell_value(ID+3, 28) != ''):
                absence = float(data.cell_value(ID+3, 28))
            elif data.cell_value(ID+2, 28) != '' and data.cell_value(ID+3, 28) == '':
                absence = float(data.cell_value(ID+2, 28))
            elif data.cell_value(ID+3, 28) == '' and data.cell_value(ID+3, 28) == '':
                absence = 0
            else: 
                absence = float(data.cell_value(ID+2, 28)) + float(data.cell_value(ID+3, 28))    

            if (data.cell_value(ID+2, 29) == '' and data.cell_value(ID+3, 29) != ''):
                normalOT = float(data.cell_value(ID+3, 29))
            elif data.cell_value(ID+2, 29) != '' and data.cell_value(ID+3, 29) == '':
                normalOT = float(data.cell_value(ID+2, 29))
            elif data.cell_value(ID+3, 29) == '' and data.cell_value(ID+3, 29) == '':
                normalOT = 0
            else: 
                normalOT = float(data.cell_value(ID+2, 29)) + float(data.cell_value(ID+3, 29))    

            if (data.cell_value(ID+2, 30) == '' and data.cell_value(ID+3, 30) != ''):
                weekendOT = float(data.cell_value(ID+3, 30))
            elif data.cell_value(ID+2, 30) != '' and data.cell_value(ID+3, 30) == '':
                weekendOT = float(data.cell_value(ID+2, 30))
            elif data.cell_value(ID+3, 30) == '' and data.cell_value(ID+3, 30) == '':
                weekendOT = 0
            else: 
                weekendOT = float(data.cell_value(ID+2, 30)) + float(data.cell_value(ID+3, 30))    

            if (data.cell_value(ID+2, 31) == '' and data.cell_value(ID+3, 31) != ''):
                holidayOT = float(data.cell_value(ID+3, 31))
            elif data.cell_value(ID+2, 31) != '' and data.cell_value(ID+3, 31) == '':
                holidayOT = float(data.cell_value(ID+2, 31))
            elif data.cell_value(ID+3, 31) == '' and data.cell_value(ID+3, 31) == '':
                holidayOT = 0
            else: 
                holidayOT = float(data.cell_value(ID+2, 31)) + float(data.cell_value(ID+3, 31))   

            if (data.cell_value(ID+2, 32) == '' and data.cell_value(ID+3, 32) != ''):
                OT1 = float(data.cell_value(ID+3, 32))
            elif data.cell_value(ID+2, 32) != '' and data.cell_value(ID+3, 32) == '':
                OT1 = float(data.cell_value(ID+2, 32))
            elif data.cell_value(ID+3, 32) == '' and data.cell_value(ID+3, 32) == '':
                OT1 = 0
            else: 
                OT1 = float(data.cell_value(ID+2, 32)) + float(data.cell_value(ID+3, 32))  

            if (data.cell_value(ID+2, 33) == '' and data.cell_value(ID+3, 33) != ''):
                OT2 = float(data.cell_value(ID+3, 33))
            elif data.cell_value(ID+2, 33) != '' and data.cell_value(ID+3, 33) == '':
                OT2 = float(data.cell_value(ID+2, 33))
            elif data.cell_value(ID+3, 33) == '' and data.cell_value(ID+3, 33) == '':
                OT2 = 0
            else: 
                OT2 = float(data.cell_value(ID+2, 33)) + float(data.cell_value(ID+3, 33))  

            if (data.cell_value(ID+2, 34) == '' and data.cell_value(ID+3, 34) != ''):
                OT3 = float(data.cell_value(ID+3, 34))
            elif data.cell_value(ID+2, 34) != '' and data.cell_value(ID+3, 34) == '':
                OT3 = float(data.cell_value(ID+2, 34))
            elif data.cell_value(ID+3, 34) == '' and data.cell_value(ID+3, 34) == '':
                OT3 = 0
            else: 
                OT3 = float(data.cell_value(ID+2, 34)) + float(data.cell_value(ID+3, 34))  

            if (data.cell_value(ID+2, 35) == '' and data.cell_value(ID+3, 35) != ''):
                xinLamThem = float(data.cell_value(ID+3, 35))
            elif data.cell_value(ID+2, 35) != '' and data.cell_value(ID+3, 35) == '':
                xinLamThem = float(data.cell_value(ID+2, 35))
            elif data.cell_value(ID+3, 35) == '' and data.cell_value(ID+3, 35) == '':
                OT3 = 0
            else: 
                xinLamThem = float(data.cell_value(ID+2, 35)) + float(data.cell_value(ID+3, 35))  

            w_sheet.write(ID+3, 26, lateIn)
            w_sheet.write(ID+3, 27, earlyOut)
            w_sheet.write(ID+3, 28, absence)
            w_sheet.write(ID+3, 29, normalOT)
            w_sheet.write(ID+3, 30, weekendOT)
            w_sheet.write(ID+3, 31, holidayOT)
            w_sheet.write(ID+3, 32, OT1)
            w_sheet.write(ID+3, 33, OT2)
            w_sheet.write(ID+3, 34, OT3)
            w_sheet.write(ID+3, 35, xinLamThem)

    wb.save('../cham-cong/convert/baocao.xlsx')


    # =========================== Get data =====================
    print("Get data va chuan bi bao cao")
    chamcong = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)

    OT_approve = xlrd.open_workbook('../cham-cong/convert/OT_convert.xlsx')
    dataOT_approve = OT_approve.sheet_by_index(0)

    #Tạo template báo cáo

    baocao = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    mod_baocao = copy(baocao)
    w_sheet_baocao = mod_baocao.get_sheet(0)

    colect_id = []
    colect_date = []
    for id in range(data.nrows-3):
        colect_id.append(data.cell_value(id+3, 0))
        colect_date.append(data.cell_value(id+3, 3))
    c = collections.Counter(colect_id)
    d = collections.Counter(colect_date)
    ID_baocao = c.keys()
    date_baocao = d.keys()
    colen2 = 0
    for y in date_baocao:
        w_sheet_baocao.write(5, colen2+6, y)
        colen2 = colen2+2

    nhanvien = xlrd.open_workbook(namenhanvien)
    sh_nhanvien = nhanvien.sheet_by_index(0)

    colen = 0
    for z in ID_baocao:
        for j in range(data.nrows -3):
            if z == sh_nhanvien.cell_value(j+3, 0):
                w_sheet_baocao.write(colen+7, 0, sh_nhanvien.cell_value(j+3, 0))
                w_sheet_baocao.write(colen+7, 1, sh_nhanvien.cell_value(j+3, 1))
                w_sheet_baocao.write(colen+7, 2, sh_nhanvien.cell_value(j+3, 3))
                w_sheet_baocao.write(colen+7, 3, sh_nhanvien.cell_value(j+3, 2))
                w_sheet_baocao.write(colen+7, 4, sh_nhanvien.cell_value(j+3, 6))
                w_sheet_baocao.write(colen+7, 5, "")
                colen = colen+1
                break
            else:
                continue
  

    for o in range(1,1):
        w_sheet_baocao.write(5, o+67, o)

    mod_baocao.save('../cham-cong/convert/baocao.xlsx')

    baocao = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    mod_baocao = copy(baocao)
    w_sheet_baocao = mod_baocao.get_sheet(0)




    baocao_1 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)
    mod_day_baocao = copy(baocao_1)
    w_sheet_baocao_day = mod_day_baocao.get_sheet(0)

    # =========================== Mã hoá ca =====================
    print("Ma hoa ca va OT")


    for m in tqdm(range(data.nrows-3)):
        if("Toi" in str(data.cell_value(m + 3, 5)) and data.cell_value(m + 4, 12) != "None" and data.cell_value(m + 3, 12) != "None" and data.cell_value(m + 4, 11) == "None" and ("Sang" in str(data.cell_value(m + 4, 5)))):
            if("Cuoi tuan" in str(data.cell_value(m + 4, 5))):
                w_sheet.write(m+4, 11, data.cell_value(m + 3, 12))
                temp = data.cell_value(m + 4, 30)
                w_sheet.write(m+4, 30, float(temp) + 2)

            else: 
                if(float(data.cell_value(m + 3, 25)) >= 5):
                    w_sheet.write(m+4, 11, data.cell_value(m + 3, 12))
                    temp = data.cell_value(m + 4, 29)
                    w_sheet.write(m+4, 29, float(temp) + 2)

    wb.save('../cham-cong/convert/baocao.xlsx')


    chamcong = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)
    
    # =========================== Duyệt OT =====================
    print("Duyet OT")
    for j in tqdm(range(data.nrows-3)):
        if dataOT_approve.nrows == 0:
            w_sheet.write(j+3, 35,myround(float( data.cell_value(j+3, 29))+float( data.cell_value(j+3, 30))+float( data.cell_value(j+3, 31))))
        else:
            for i in range(dataOT_approve.nrows):
                if  dataOT_approve.cell_value(i, 0) == data.cell_value(j+3, 0) and dataOT_approve.cell_value(i, 2) == data.cell_value(j+3, 3):
                    w_sheet.write(j+3, 35, myround(float( data.cell_value(j+3, 29))+float( data.cell_value(j+3, 32))))
                    break
                else:
                    w_sheet.write(j+3, 35, myround(float( data.cell_value(j+3, 29))+float( data.cell_value(j+3, 30))+float( data.cell_value(j+3, 31))))
    
    for j in tqdm(range(data.nrows-3)):
        otTime = (float( data.cell_value(j+3, 29)) + float( data.cell_value(j+3, 30)) + float( data.cell_value(j+3, 31)))
        if ((otTime) <= (otTime + float( data.cell_value(j+3, 32)) - 2)):
            w_sheet.write(j+3, 35,"RR")


    for m in tqdm(range(data.nrows-3)):

        #Kiem tra ca
        if(float(data.cell_value(m+3, 25))>=5):
            if(data.cell_value(m+3, 5) == "Sản xuất Sáng" or data.cell_value(m+3, 5) == "Bảo trì Sáng"):
                w_sheet.write(m+3, 36, "A")
            elif(data.cell_value(m+3, 5) == "Sản xuất Tối" or data.cell_value(m+3, 5) == "Bảo trì Tối"):
                w_sheet.write(m+3, 36, "C")
            elif(data.cell_value(m+3, 5) == "Ca Chiều"):
                w_sheet.write(m+3, 36, "B")
            elif("Hành Chính" in data.cell_value(m+3, 5)):
                w_sheet.write(m+3, 36, "D")
        elif(float(data.cell_value(m+3, 25))<5 or float(data.cell_value(m+3, 25))>=3):
            if(data.cell_value(m+3, 5) == "Sản xuất Sáng" or data.cell_value(m+3, 5) == "Bảo trì Sáng"):
                w_sheet.write(m+3, 36, "RR5")
            elif(data.cell_value(m+3, 5) == "Sản xuất Tối" or data.cell_value(m+3, 5) == "Bảo trì Tối"):
                w_sheet.write(m+3, 36, "RR5")
            elif(data.cell_value(m+3, 5) == "Ca Chiều"):
                w_sheet.write(m+3, 36, "RR5")
            elif("Hành Chính" in data.cell_value(m+3, 5)):
                w_sheet.write(m+3, 36, "RR5")
        else:
            w_sheet.write(m+3, 36, "RR")
 
         #Kiem tra quen cham cong
        if (data.cell_value(m+3, 11) == "None" or data.cell_value(m+3, 12) == "None"):
            w_sheet.write(m+3, 36, "RR")
 
        #Kiem tra thu 7
        if(datetime.strptime(data.cell_value(m+3, 3), "%Y-%m-%d").weekday()==5 and data.cell_value(m+3, 22) == "Gián Tiếp"):
            if(float(data.cell_value(m+3, 25))<5):
                w_sheet.write(m+3, 36, "nt7")
            else:
                w_sheet.write(m+3, 36, "D")

        #Kiem tra chu nhat
        if(datetime.strptime(data.cell_value(m+3, 3), "%Y-%m-%d").weekday()==6):
            if(float(data.cell_value(m+3, 30)) > 1 and data.cell_value(m+3, 5) == ""):
                w_sheet.write(m+3, 36, "CN")
            else:
                w_sheet.write(m+3, 36, "")
 
    wb.save('../cham-cong/convert/baocao.xlsx')



    # =========================== Chuyển dữ liệu sang report =============================================
    chamcong = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)


    # # Chuyển ngày
    print("Chuyen du lieu vao bao cao")
    oi = len(date_baocao)*2
    for i in tqdm(range(data_baocao.nrows-7)):
        for j in range(data.nrows-3):
            if data_baocao.cell_value(i+7, 0) == data.cell_value(j+3, 0):
                for k in range(0,oi,2):
                    if data_baocao.cell_value(5, k+6) == data.cell_value(j+3, 3):
                            w_sheet_baocao_day.write(i+7,  k+6, data.cell_value(j+3, 36))
                            w_sheet_baocao_day.write(i+7,  k+7, data.cell_value(j+3, 35))
    mod_day_baocao.save('../cham-cong/convert/baocao.xlsx')

    print("Chuyen du lieu vao bao cao vi pham")
    oi = len(date_baocao)*2
    for i in tqdm(range(data_baocao.nrows-7)):
        for j in range(data.nrows-3):
            if data_baocao.cell_value(i+7, 0) == data.cell_value(j+3, 0):
                for k in range(0,oi,2):
                    if data_baocao.cell_value(5, k+6) == data.cell_value(j+3, 3):
                        if data.cell_value(j+3, 11) == "None":
                            w_sheet_baocao.write(i+7,  k+6, "QCC")
                        elif data.cell_value(j+3, 26) != "":
                            kll = float(data.cell_value(j+3, 26))
                            w_sheet_baocao.write(i+7,  k+6, kll)
                        if data.cell_value(j+3, 12) == "None":
                            w_sheet_baocao.write(i+7,  k+7, "QCC")
                        elif data.cell_value(j+3, 27) != "":
                            kl = float(data.cell_value(j+3, 27))
                            w_sheet_baocao.write(i+7,  k+7, kl)
                        if data.cell_value(j+3, 11) == "None" and data.cell_value(j+3, 12) == "None":
                            w_sheet_baocao.write(i+7,  k+6, "Nghi")
                            w_sheet_baocao.write(i+7,  k+7, "")
                        if data.cell_value(j+3, 11) == "None" and data.cell_value(j+3, 12) == "None" and data.cell_value(j+3, 5) == "":
                            w_sheet_baocao.write(i+7,  k+6, "")
                    
    mod_baocao.save('../cham-cong/convert/baocaovipham.xlsx')

    # =========================== xử lý file mở k được =====================
    print("Report")
    #Data
    all_rows_data = []
    for row in range(data.nrows):
        curr_row = []
        for col in range(data.ncols):
            curr_row.append(data.cell_value(row, col))
        all_rows_data.append(curr_row)

    chamcong1 = xlsxwriter.Workbook("../cham-cong/convert/data1.xlsx")
    data1 = chamcong1.add_worksheet()

    for row in range(len(all_rows_data)):
        for col in range(len(all_rows_data[0])):
            data1.write(row, col, all_rows_data[row][col])
    chamcong1.close()

    baocao_2 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_2.sheet_by_index(0)

    # BAO CAO
    all_rows_baocao = []
    for row in range(data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    baocao2 = xlsxwriter.Workbook('../cham-cong/convert/baocao2.xlsx')
    data2 = baocao2.add_worksheet()

    for row in range(len(all_rows_baocao)):
        for col in range(len(all_rows_baocao[0])):
            data2.write(row, col, all_rows_baocao[row][col])
    baocao2.close()


    baocao_1 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)


    # ====================================Chuyen du lieu vao report==========================
    baocao_2 = xlrd.open_workbook('../cham-cong/convert/baocao.xlsx')
    data_baocao = baocao_2.sheet_by_index(0)
    all_rows_baocao = []
    
    for row in range(7 ,data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    data_convert = openpyxl.load_workbook('../cham-cong/template/Template_report.xlsx')
    sheet_name_data_convert = data_convert.sheetnames[0]
    sh_data_convert = data_convert[sheet_name_data_convert]
    sh_data_convert.cell(6, 4).value = text_thang*1
    sh_data_convert.cell(6, 6).value = text_nam*1


    for row in tqdm(range(1, len(all_rows_baocao)+1)):
        for col in range(1, len(all_rows_baocao[0])+1):
            sh_data_convert.cell(row+10, col).value = all_rows_baocao[row-1][col-1]
    data_convert.save("../cham-cong/report/chamcong"+"_thang"+text_thang+"_nam"+text_nam +".xlsx")

    # BAO CAO Vi PHAM
    
    baocao_1 = xlrd.open_workbook('../cham-cong/convert/baocaovipham.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)

    all_rows_baocao = []
    for row in range(data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    baocao1 = xlsxwriter.Workbook("../cham-cong/convert/baocaovipham.xlsx")
    data2 = baocao1.add_worksheet()

    for row in range(len(all_rows_baocao)):
        for col in range(len(all_rows_baocao[0])):
            data2.write(row, col, all_rows_baocao[row][col])
    baocao1.close()

    # ====================================Chuyen du lieu vao report vi pham==========================

    all_rows_baocao = []
    for row in range(7 ,data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    data_convert = openpyxl.load_workbook('../cham-cong/template/Template_report_vipham.xlsx')
    sheet_name_data_convert = data_convert.sheetnames[0]
    sh_data_convert = data_convert[sheet_name_data_convert]
    sh_data_convert.cell(6, 4).value = text_thang*1
    sh_data_convert.cell(6, 6).value = text_nam*1


    for row in tqdm(range(1, len(all_rows_baocao)+1)):
        for col in range(1, len(all_rows_baocao[0])+1):
            sh_data_convert.cell(row+10, col).value = all_rows_baocao[row-1][col-1]
    data_convert.save("../cham-cong/report/baocao_vipham"+"_thang"+text_thang+"_nam"+text_nam + ".xlsx")
    print("done")              