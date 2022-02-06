import xlwt
import xlrd
from xlutils.copy import copy
import xlsxwriter
import collections
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm, trange
from datetime import date, datetime
def xuly(namedata, nameOT,namenhanvien, valueyear, valuemounth):
    def myround(x, base=0.5):
        return base * round(float(x) / base)
    # =========================== convert OT =====================
    print("Chuan bi du lieu")
    dataOT = xlrd.open_workbook(nameOT)
    ot = dataOT.sheet_by_index(0)
    ot_convert = xlsxwriter.Workbook('OT_convert.xlsx')
    add_sheet = ot_convert.add_worksheet()
    for i in range(ot.nrows-3):
        id = ot.cell_value(i+3, 0)
        date = ot.cell_value(i+3, 4)
        date1 =ot.cell_value(i+3, 5)
        x =(datetime.strptime(date,"%Y-%m-%d %H:%M:%S"))
        y =(datetime.strptime(date1,"%Y-%m-%d %H:%M:%S"))
        timeOT = y - x
        hh, mm , ss = map(int, str(timeOT).split(':'))
        ot3 = hh + mm/60
        d = x.strftime("%Y-%m-%d")
        add_sheet.write(i,0, str(id))
        add_sheet.write(i,1, ot3)
        
        add_sheet.write(i,2, d)
    ot_convert.close()

    # =========================== Xoá data =====================

    baocao_del = xlrd.open_workbook('baocao.xlsx')
    data_baocao_del = baocao_del.sheet_by_index(0)

    all_rows_baocao_del = []
    for row in tqdm(range(data_baocao_del.nrows)):
        curr_row = []
        for col in range(data_baocao_del.ncols):
            curr_row.append(data_baocao_del.cell_value(row, col))
        all_rows_baocao_del.append(curr_row)

    delete_baocao = xlsxwriter.Workbook('baocao.xlsx')
    delete = delete_baocao.add_worksheet()

    for row in range(len(all_rows_baocao_del)):
        for col in range(len(all_rows_baocao_del[0])):
            delete.write(row, col, "")
    delete_baocao.close()


    # =========================== Get data =====================
    print("Get data va chuan bi bao cao")
    chamcong = xlrd.open_workbook(namedata)
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)

    OT_approve = xlrd.open_workbook('OT_convert.xlsx')
    dataOT_approve = OT_approve.sheet_by_index(0)

    #Tạo template báo cáo

    baocao = xlrd.open_workbook('baocao.xlsx')
    mod_baocao = copy(baocao)
    w_sheet_baocao = mod_baocao.get_sheet(0)

    colect_id = []
    colect_date = []
    for id in range(data.nrows-3):
        colect_id.append(data.cell_value(id+3, 0))
        colect_date.append(data.cell_value(id+3, 3))
    c = collections.Counter(colect_id)
    d = collections.Counter(colect_date)
    ID_baocao = c.keys()
    date_baocao = d.keys()

    colen2 = 0
    for y in date_baocao:
        w_sheet_baocao.write(5, colen2+6, y)
        colen2 = colen2+2

    nhanvien = xlrd.open_workbook(namenhanvien)
    sh_nhanvien = nhanvien.sheet_by_index(0)

    colen = 0
    for z in ID_baocao:
        for j in range(data.nrows -3):
            if z == sh_nhanvien.cell_value(j+3, 0):
                w_sheet_baocao.write(colen+7, 0, sh_nhanvien.cell_value(j+3, 0))
                w_sheet_baocao.write(colen+7, 1, sh_nhanvien.cell_value(j+3, 1))
                w_sheet_baocao.write(colen+7, 2, sh_nhanvien.cell_value(j+3, 3))
                w_sheet_baocao.write(colen+7, 3, sh_nhanvien.cell_value(j+3, 2))
                w_sheet_baocao.write(colen+7, 4, sh_nhanvien.cell_value(j+3, 6))
                w_sheet_baocao.write(colen+7, 5, "")
                colen = colen+1
                break
  

    for o in range(1,1):
        w_sheet_baocao.write(5, o+67, o)

    mod_baocao.save('baocao.xlsx')

    baocao = xlrd.open_workbook('baocao.xlsx')
    mod_baocao = copy(baocao)
    w_sheet_baocao = mod_baocao.get_sheet(0)



    # =========================== Duyệt OT =====================
    print("Duyet OT")
    for j in tqdm(range(data.nrows-3)):
        if dataOT_approve.nrows == 0:
            w_sheet.write(j+3, 35,myround(float( data.cell_value(j+3, 29))+float( data.cell_value(j+3, 30))+float( data.cell_value(j+3, 31))))
        else:
            for i in range(dataOT_approve.nrows):
                if  dataOT_approve.cell_value(i, 0) == data.cell_value(j+3, 0) and dataOT_approve.cell_value(i, 2) == data.cell_value(j+3, 3):
                    w_sheet.write(j+3, 35, myround(float( data.cell_value(j+3, 29))+float( data.cell_value(j+3, 32))))
                    break
                else:
                    w_sheet.write(j+3, 35, myround(float( data.cell_value(j+3, 29))+float( data.cell_value(j+3, 30))+float( data.cell_value(j+3, 31))))
    
    for j in tqdm(range(data.nrows-3)):
        otTime = (float( data.cell_value(j+3, 29)) + float( data.cell_value(j+3, 30)) + float( data.cell_value(j+3, 31)))
        if ((otTime) <= (otTime + float( data.cell_value(j+3, 32)) - 2)):
            w_sheet.write(j+3, 35,"RR")




    baocao_1 = xlrd.open_workbook('baocao.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)
    mod_day_baocao = copy(baocao_1)
    w_sheet_baocao_day = mod_day_baocao.get_sheet(0)

    # =========================== Mã hoá ca =====================
    print("Ma hoa ca va OT")
    for m in tqdm(range(data.nrows-3)):

        #Kiem tra ca
        if(float(data.cell_value(m+3, 25))>=5):
            if(data.cell_value(m+3, 5) == "San xuat Sang"):
                w_sheet.write(m+3, 36, "A")
            elif(data.cell_value(m+3, 5) == "San xuat Toi"):
                w_sheet.write(m+3, 36, "C")
            elif(data.cell_value(m+3, 5) == "San xuat Ca C"):
                w_sheet.write(m+3, 36, "B")
            elif(data.cell_value(m+3, 5) == "Ca Hanh Chính"):
                w_sheet.write(m+3, 36, "D")
        elif(float(data.cell_value(m+3, 25))<5 or float(data.cell_value(m+3, 25))>=3):
            if(data.cell_value(m+3, 5) == "San xuat Sang"):
                w_sheet.write(m+3, 36, "RR5")
            elif(data.cell_value(m+3, 5) == "San xuat Toi"):
                w_sheet.write(m+3, 36, "RR5")
            elif(data.cell_value(m+3, 5) == "San xuat Ca C"):
                w_sheet.write(m+3, 36, "RR5")
            elif(data.cell_value(m+3, 5) == "Ca Hanh Chính"):
                w_sheet.write(m+3, 36, "RR5")
        else:
            w_sheet.write(m+3, 36, "RR")
 
         #Kiem tra quen cham cong
        if (data.cell_value(m+3, 11) == "None" or data.cell_value(m+3, 12) == "None"):
            w_sheet.write(m+3, 36, "RR")
 
        #Kiem tra thu 7
        if(datetime.strptime(data.cell_value(m+3, 3), "%Y-%m-%d").weekday()==5 and data.cell_value(m+3, 22) == "Gián Tiếp"):
            if(float(data.cell_value(m+3, 25))<5):
                w_sheet.write(m+3, 36, "nt7")
            else:
                w_sheet.write(m+3, 36, "D")

        #Kiem tra chu nhat
        if(datetime.strptime(data.cell_value(m+3, 3), "%Y-%m-%d").weekday()==6):
            if(float(data.cell_value(m+3, 30)) > 1 ):
                w_sheet.write(m+3, 36, "CN")
            else:
                w_sheet.write(m+3, 36, "")
 
    wb.save('baocao.xlsx')



    # =========================== Chuyển dữ liệu sang report =============================================
    chamcong = xlrd.open_workbook('baocao.xlsx')
    data = chamcong.sheet_by_index(0)
    wb = copy(chamcong)
    w_sheet = wb.get_sheet(0)


    # # Chuyển ngày
    print("Chuyen du lieu vao bao cao")
    oi = len(date_baocao)*2
    for i in tqdm(range(data_baocao.nrows-7)):
        for j in range(data.nrows-3):
            if data_baocao.cell_value(i+7, 0) == data.cell_value(j+3, 0):
                for k in range(0,oi,2):
                    if data_baocao.cell_value(5, k+6) == data.cell_value(j+3, 3):
                            w_sheet_baocao_day.write(i+7,  k+6, data.cell_value(j+3, 36))
                            w_sheet_baocao_day.write(i+7,  k+7, data.cell_value(j+3, 35))
    mod_day_baocao.save('baocao.xlsx')

    # =========================== xử lý file mở k được =====================
    print("Report")
    #Data
    all_rows_data = []
    for row in range(data.nrows):
        curr_row = []
        for col in range(data.ncols):
            curr_row.append(data.cell_value(row, col))
        all_rows_data.append(curr_row)

    chamcong1 = xlsxwriter.Workbook('data1.xlsx')
    data1 = chamcong1.add_worksheet()

    for row in range(len(all_rows_data)):
        for col in range(len(all_rows_data[0])):
            data1.write(row, col, all_rows_data[row][col])
    chamcong1.close()

    baocao_2 = xlrd.open_workbook('baocao.xlsx')
    data_baocao = baocao_2.sheet_by_index(0)

    # BAO CAO
    all_rows_baocao = []
    for row in range(data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    baocao1 = xlsxwriter.Workbook('baocao1.xlsx')
    data2 = baocao1.add_worksheet()

    for row in range(len(all_rows_baocao)):
        for col in range(len(all_rows_baocao[0])):
            data2.write(row, col, all_rows_baocao[row][col])
    baocao1.close()


    baocao_1 = xlrd.open_workbook('baocao.xlsx')
    data_baocao = baocao_1.sheet_by_index(0)




    # ====================================Chuyen du lieu vao report

    all_rows_baocao = []
    for row in range(7 ,data_baocao.nrows):
        curr_row = []
        for col in range(data_baocao.ncols):
            curr_row.append(data_baocao.cell_value(row, col))
        all_rows_baocao.append(curr_row)

    data_convert = openpyxl.load_workbook('Template_report.xlsx')
    sheet_name_data_convert = data_convert.sheetnames[0]
    sh_data_convert = data_convert[sheet_name_data_convert]
    sh_data_convert.cell(6, 4).value = "1"
    sh_data_convert.cell(6, 6).value = "2022"


    for row in tqdm(range(1, len(all_rows_baocao)+1)):
        for col in range(1, len(all_rows_baocao[0])+1):
            sh_data_convert.cell(row+10, col).value = all_rows_baocao[row-1][col-1]
    data_convert.save("report1.xlsx")
    print("done")              